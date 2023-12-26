import openpyxl
import sys
import mariadb
from datetime import datetime


def conectar_mariadb(usuario, contra, host_ip, puerto, db):
    try:
        conn = mariadb.connect(
            user=usuario,
            password=contra,
            host=host_ip,
            port=puerto,
            database =db
        )
    except mariadb.Error as e:  
        print(f"Error connecting to MariaDB Platform: {e}")
        sys.exit(1)
    return conn


def conexion (conn, lista):
    cur = conn.cursor()
    print("Aqui")
    fecha_texto = lista[1]  # Obtiene la cadena de texto que representa la fecha
    fecha_datetime = datetime.strptime(fecha_texto, '%Y-%m-%d %H:%M:%S')  # Convierte la cadena a datetime
    fecha_date = fecha_datetime.date()  # Obtiene la parte de fecha (sin la hora)
    print("aca")

    cur.execute(
        """INSERT INTO despachos_cesfam(numero_de_receta, fecha_de_movimiento, usuario, establecimiento, sexo,
        escolaridad, pais_de_origen, pueblo_originario, sector, discapacidad, prevision, tipo_receta, articulo,
        posologia_de_la_descripcion, observacion, dosis, cada, unidad_de_tiempo, por, unidad_de_tiempo2, estado_de_receta,
        cantidad_de_despacho, cantidad_sin_despacho, diagnostico, codigo_cie10, edad_años, profesional, sector_inscripcion)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
            lista[0],
            fecha_date,
            lista[2],
            lista[3],
            lista[4],
            lista[5],
            lista[6],
            lista[7],
            lista[8],
            lista[9],
            lista[10],
            lista[11],
            lista[12],
            lista[13],
            lista[14],
            lista[15],
            lista[16],
            lista[17],
            lista[18],
            lista[19],
            lista[20],
            lista[21],
            lista[22],
            lista[23],
            lista[24],
            lista[25],
            lista[26],
            lista[27]
        ),
    )
    conn.commit()
   



# Conectar a mariadb
conn = conectar_mariadb("root","Lom1smo1!","172.30.96.1",3306, "cesfam")

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Informe_Rayen.xlsx")

for sheet_index in range(0, len(dataframe.sheetnames)): #cantidad de paginas dentro del archivo
    current_sheet = dataframe.worksheets[sheet_index] # pagina en la que trabajaremos
    for row in range(2, current_sheet.max_row + 1): # cantidad de filas
        lista = []
        for col in current_sheet.iter_cols(1, current_sheet.max_column): #cantidad de columnas del archivo
            lista.append(str(col[row - 1].value))  # Resta 1 al índice de fila para coincidir con el índice de Python 
        print(lista[1])  
        conexion(conn, lista)

conn.close()

